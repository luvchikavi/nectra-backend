# in backend/apps/projects/views.py

from rest_framework import viewsets, status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django.db.models import Sum, Count
from django.utils import timezone
from decimal import Decimal
import openpyxl
from datetime import datetime

from .models import Project, ProjectDataInputs, BankTransaction, ConstructionProgress, ConstructionProgressSnapshot, EquityDeposit, ProjectDocument
from .serializers import (
    ProjectSerializer,
    BankTransactionSerializer,
    BankStatementUploadSerializer,
    BankTransactionApprovalSerializer,
    BankTransactionBulkApprovalSerializer,
    ConstructionProgressSerializer,
    ConstructionProgressSnapshotSerializer,
    EquityDepositSerializer,
    ProjectDocumentSerializer,
    ProjectDocumentUploadSerializer
)


class ProjectViewSet(viewsets.ModelViewSet):
    queryset = Project.objects.all()
    serializer_class = ProjectSerializer

    def get_queryset(self):
        """By default, return only active projects. Use ?include_deleted=true to see all."""
        # Use select_related to optimize apartments_count serializer method
        queryset = Project.objects.select_related('data_inputs').prefetch_related('apartments')
        include_deleted = self.request.query_params.get('include_deleted', 'false').lower() == 'true'
        if not include_deleted:
            queryset = queryset.filter(is_active=True)
        return queryset

    @action(detail=True, methods=['post'], url_path='soft-delete')
    def soft_delete(self, request, pk=None):
        """Soft delete a project - marks as inactive but keeps all data"""
        project = self.get_object()

        # Check if project has any data that should prevent deletion
        warnings = []

        # Check for associated invoices
        if hasattr(project, 'contractor_invoices') and project.contractor_invoices.exists():
            invoice_count = project.contractor_invoices.count()
            warnings.append(f'הפרויקט מכיל {invoice_count} חשבוניות / Project has {invoice_count} invoices')

        # Check for bank transactions
        if hasattr(project, 'bank_transactions') and project.bank_transactions.exists():
            tx_count = project.bank_transactions.count()
            warnings.append(f'הפרויקט מכיל {tx_count} תנועות בנק / Project has {tx_count} bank transactions')

        # Check for apartments (sold ones are critical)
        if hasattr(project, 'apartments'):
            apartments = project.apartments.all()
            sold_count = apartments.filter(unit_status='SOLD').count()
            if sold_count > 0:
                warnings.append(f'הפרויקט מכיל {sold_count} דירות שנמכרו / Project has {sold_count} sold apartments')

        # Perform soft delete
        project.soft_delete()

        return Response({
            'status': 'success',
            'message': 'הפרויקט הועבר לפח / Project moved to trash',
            'project_id': project.id,
            'project_name': project.project_name,
            'warnings': warnings,
            'can_restore': True
        })

    @action(detail=True, methods=['post'], url_path='restore')
    def restore(self, request, pk=None):
        """Restore a soft-deleted project"""
        # Override queryset to include deleted projects for this action
        project = Project.objects.get(pk=pk)

        if project.is_active:
            return Response({
                'status': 'error',
                'message': 'הפרויקט כבר פעיל / Project is already active'
            }, status=status.HTTP_400_BAD_REQUEST)

        project.restore()

        return Response({
            'status': 'success',
            'message': 'הפרויקט שוחזר בהצלחה / Project restored successfully',
            'project_id': project.id,
            'project_name': project.project_name
        })

    @action(detail=False, methods=['get'], url_path='deleted')
    def deleted_projects(self, request):
        """Get list of soft-deleted projects"""
        deleted = Project.objects.filter(is_active=False)
        serializer = self.get_serializer(deleted, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'], url_path='month-info')
    def month_info(self, request, pk=None):
        """Get project month information (current month X of total Y months)"""
        project = self.get_object()

        try:
            data_inputs = project.data_inputs
            dates = data_inputs.dates or {}

            construction_start_date_str = dates.get('construction_start_date')
            construction_duration_months = dates.get('construction_duration_months', 36)

            if construction_start_date_str:
                from datetime import datetime
                from dateutil.relativedelta import relativedelta

                construction_start_date = datetime.strptime(construction_start_date_str, '%Y-%m-%d').date()
                today = datetime.now().date()

                # Calculate how many months have passed since construction start
                if today < construction_start_date:
                    current_month = 0
                else:
                    diff = relativedelta(today, construction_start_date)
                    current_month = diff.years * 12 + diff.months + 1  # +1 because month 1 is the first month

                # Cap at total duration
                if current_month > construction_duration_months:
                    current_month = construction_duration_months

                return Response({
                    'current_month': current_month,
                    'total_months': construction_duration_months,
                    'construction_start_date': construction_start_date_str,
                    'display_text': f'חודש {current_month} מתוך {construction_duration_months}',
                    'display_text_en': f'Month {current_month} of {construction_duration_months}'
                })
            else:
                return Response({
                    'current_month': None,
                    'total_months': construction_duration_months,
                    'construction_start_date': None,
                    'display_text': 'לא הוגדר תאריך התחלת בנייה',
                    'display_text_en': 'Construction start date not defined'
                })

        except ProjectDataInputs.DoesNotExist:
            return Response({
                'current_month': None,
                'total_months': None,
                'construction_start_date': None,
                'display_text': 'אין נתוני פרויקט',
                'display_text_en': 'No project data'
            })


class CityChoicesView(APIView):
    """
    An API view to provide the list of city choices for frontend forms.
    """
    def get(self, request, *args, **kwargs):
        formatted_cities = [
            {"value": choice[0], "label": choice[1]}
            for choice in Project.CITY_CHOICES
        ]
        return Response(formatted_cities)


class DashboardStatsView(APIView):
    """
    An API view to provide dashboard statistics across all projects.
    """
    def get(self, request, *args, **kwargs):
        from apps.sales.models import SalesTransaction, ApartmentInventory
        from django.db.models import Count, Sum, Avg

        # Get total projects
        total_projects = Project.objects.count()

        # Get total apartments and sold count - include Section 7 data (optimized)
        try:
            total_apartments = ApartmentInventory.objects.count()
            sold_apartments = ApartmentInventory.objects.filter(unit_status='SOLD').count()

            # Get project IDs that have ApartmentInventory records
            projects_with_apartments = set(
                ApartmentInventory.objects.values_list('project_id', flat=True).distinct()
            )

            # Add Section 7 data for projects without ApartmentInventory records
            section7_total = 0
            section7_sold = 0

            # Use select_related to fetch data_inputs in one query
            for project in Project.objects.select_related('data_inputs').only('id'):
                if project.id not in projects_with_apartments:
                    try:
                        revenue_forecast = project.data_inputs.revenue_forecast or {}
                        revenue_residential = revenue_forecast.get('revenue_residential', [])
                        if revenue_residential:
                            section7_total += len(revenue_residential)
                            for unit in revenue_residential:
                                status = str(unit.get('status', '')).lower()
                                if 'נמכר' in status or 'sold' in status:
                                    section7_sold += 1
                    except:
                        pass

            total_apartments += section7_total
            sold_apartments += section7_sold
            sold_percent = round((sold_apartments / total_apartments) * 100, 1) if total_apartments > 0 else 0
        except:
            total_apartments = 0
            sold_apartments = 0
            sold_percent = 0

        # Get equity deposit stats
        total_equity = EquityDeposit.objects.aggregate(total=Sum('amount'))['total'] or Decimal('0')

        # Get construction progress stats
        try:
            progress_data = ConstructionProgress.objects.aggregate(
                avg_completion=Avg('overall_completion_percentage')
            )
            avg_construction_progress = round(float(progress_data['avg_completion'] or 0), 1)
        except:
            avg_construction_progress = 0

        # Get pending bank transactions (urgent items)
        pending_transactions = BankTransaction.objects.filter(status='PENDING').count()

        # Get sales stats
        try:
            total_sales_value = SalesTransaction.objects.aggregate(
                total=Sum('final_price')
            )['total'] or Decimal('0')
        except:
            total_sales_value = Decimal('0')

        # Get projects in construction phase
        construction_projects = Project.objects.filter(phase='CONSTRUCTION').count()
        pre_construction_projects = Project.objects.filter(phase='PRE_CONSTRUCTION').count()

        return Response({
            'total_projects': total_projects,
            'total_apartments': total_apartments,
            'sold_apartments': sold_apartments,
            'sold_percent': sold_percent,
            'total_equity': float(total_equity),
            'avg_construction_progress': avg_construction_progress,
            'pending_transactions': pending_transactions,
            'total_sales_value': float(total_sales_value),
            'construction_projects': construction_projects,
            'pre_construction_projects': pre_construction_projects,
            'urgent_items': pending_transactions  # Pending transactions as urgent items
        })


class FinancialKPIsView(APIView):
    """
    API view for company-wide financial KPIs (main dashboard).
    """
    def get(self, request, *args, **kwargs):
        from apps.sales.models import SalesTransaction, ApartmentInventory
        from django.db.models import Sum, Avg, F
        from datetime import datetime, date
        from dateutil.relativedelta import relativedelta

        # Get all active projects
        projects = Project.objects.filter(is_active=True)
        total_projects = projects.count()

        # Initialize totals
        total_revenue = Decimal('0')
        total_cost = Decimal('0')
        total_sales_value = Decimal('0')
        total_equity = Decimal('0')
        total_apartments = 0
        sold_apartments = 0
        projects_data = []

        for project in projects:
            project_revenue = Decimal('0')
            project_cost = Decimal('0')
            project_profit = Decimal('0')
            timeline_info = {}

            # Get data inputs for financial info
            try:
                data_inputs = project.data_inputs

                # Get profitability data
                profitability = data_inputs.profitability or {}
                land_value = data_inputs.land_value or {}

                # Revenue from land_value
                project_revenue = Decimal(str(land_value.get('total_income_combined_no_vat', 0) or 0))
                project_cost = Decimal(str(land_value.get('total_cost_project', 0) or 0))

                # Timeline from dates
                dates = data_inputs.dates or {}
                construction_start = dates.get('construction_start_date')
                duration_months = dates.get('construction_duration_months', 36)

                if construction_start:
                    start_date = datetime.strptime(construction_start, '%Y-%m-%d').date()
                    end_date = start_date + relativedelta(months=duration_months)
                    today = date.today()

                    if today < start_date:
                        days_remaining = (end_date - start_date).days
                        progress_percent = 0
                    elif today > end_date:
                        days_remaining = 0
                        progress_percent = 100
                    else:
                        days_remaining = (end_date - today).days
                        total_days = (end_date - start_date).days
                        elapsed_days = (today - start_date).days
                        progress_percent = round((elapsed_days / total_days) * 100, 1) if total_days > 0 else 0

                    timeline_info = {
                        'start_date': construction_start,
                        'end_date': end_date.strftime('%Y-%m-%d'),
                        'duration_months': duration_months,
                        'days_remaining': days_remaining,
                        'progress_percent': progress_percent
                    }
            except Exception:
                pass

            # Get sales data for this project - with Section 7 fallback
            try:
                project_apartments = ApartmentInventory.objects.filter(project=project)
                project_total_apts = project_apartments.count()
                project_sold_apts = project_apartments.filter(unit_status='SOLD').count()

                # If no ApartmentInventory, try Section 7 data
                if project_total_apts == 0:
                    try:
                        revenue_forecast = data_inputs.revenue_forecast or {}
                        revenue_residential = revenue_forecast.get('revenue_residential', [])
                        if revenue_residential and len(revenue_residential) > 0:
                            project_total_apts = len(revenue_residential)
                            for unit in revenue_residential:
                                status = str(unit.get('status', '')).lower()
                                if 'נמכר' in status or 'sold' in status:
                                    project_sold_apts += 1
                    except:
                        pass

                total_apartments += project_total_apts
                sold_apartments += project_sold_apts

                # Get actual sales value
                project_sales = SalesTransaction.objects.filter(
                    apartment__project=project
                ).aggregate(total=Sum('final_price'))['total'] or Decimal('0')
                total_sales_value += project_sales

                # If no data_inputs revenue, use sales value
                if project_revenue == 0 and project_sales > 0:
                    project_revenue = project_sales
            except Exception:
                project_total_apts = 0
                project_sold_apts = 0

            # Get equity deposits for this project
            try:
                project_equity = EquityDeposit.objects.filter(
                    project=project
                ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
                total_equity += project_equity
            except Exception:
                project_equity = Decimal('0')

            # Get construction progress
            try:
                construction = ConstructionProgress.objects.get(project=project)
                construction_percent = float(construction.overall_completion_percentage or 0)
            except ConstructionProgress.DoesNotExist:
                construction_percent = 0

            project_profit = project_revenue - project_cost

            total_revenue += project_revenue
            total_cost += project_cost

            projects_data.append({
                'id': project.id,
                'name': project.project_name,
                'city': project.get_city_display() if hasattr(project, 'get_city_display') else project.city,
                'revenue': float(project_revenue),
                'cost': float(project_cost),
                'profit': float(project_profit),
                'roi': round((float(project_profit) / float(project_cost) * 100), 1) if project_cost > 0 else 0,
                'profit_margin': round((float(project_profit) / float(project_revenue) * 100), 1) if project_revenue > 0 else 0,
                'apartments_total': project_total_apts,
                'apartments_sold': project_sold_apts,
                'sales_percent': round((project_sold_apts / project_total_apts * 100), 1) if project_total_apts > 0 else 0,
                'construction_progress': construction_percent,
                'equity': float(project_equity),
                'timeline': timeline_info
            })

        # Calculate company-wide KPIs
        total_profit = total_revenue - total_cost
        company_roi = round((float(total_profit) / float(total_cost) * 100), 1) if total_cost > 0 else 0
        company_profit_margin = round((float(total_profit) / float(total_revenue) * 100), 1) if total_revenue > 0 else 0
        sales_rate = round((sold_apartments / total_apartments * 100), 1) if total_apartments > 0 else 0

        # Get average construction progress
        try:
            avg_construction = ConstructionProgress.objects.aggregate(
                avg=Avg('overall_completion_percentage')
            )['avg'] or 0
        except Exception:
            avg_construction = 0

        return Response({
            'company': {
                'total_projects': total_projects,
                'total_revenue': float(total_revenue),
                'total_cost': float(total_cost),
                'total_profit': float(total_profit),
                'total_equity': float(total_equity),
                'total_sales_value': float(total_sales_value),
                'roi': company_roi,
                'profit_margin': company_profit_margin,
                'total_apartments': total_apartments,
                'sold_apartments': sold_apartments,
                'sales_rate': sales_rate,
                'avg_construction_progress': round(float(avg_construction), 1)
            },
            'projects': projects_data
        })


class ProjectFinancialKPIsView(APIView):
    """
    API view for project-specific financial KPIs.
    """
    def get(self, request, project_pk=None, *args, **kwargs):
        from apps.sales.models import SalesTransaction, ApartmentInventory, CustomerPaymentSchedule
        from django.db.models import Sum
        from datetime import datetime, date
        from dateutil.relativedelta import relativedelta

        try:
            project = Project.objects.get(pk=project_pk, is_active=True)
        except Project.DoesNotExist:
            return Response(
                {'error': 'Project not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Initialize response data
        kpis = {
            'project_id': project.id,
            'project_name': project.project_name,
            'profitability': {},
            'timeline': {},
            'sales': {},
            'construction': {},
            'cashflow': {},
            'area_metrics': {}  # New: detailed per-sqm metrics
        }

        # Get data inputs
        try:
            data_inputs = project.data_inputs

            # Profitability from land_value
            land_value = data_inputs.land_value or {}
            profitability_data = data_inputs.profitability or {}
            fixed_rates = data_inputs.fixed_rates or {}

            total_revenue = Decimal(str(land_value.get('total_income_combined_no_vat', 0) or 0))
            total_cost = Decimal(str(land_value.get('total_cost_project', 0) or 0))
            land_cost = Decimal(str(land_value.get('land_cost_balance', 0) or 0))

            profit = total_revenue - total_cost
            roi = round((float(profit) / float(total_cost) * 100), 1) if total_cost > 0 else 0
            profit_margin = round((float(profit) / float(total_revenue) * 100), 1) if total_revenue > 0 else 0
            developer_rate = fixed_rates.get('developer_profitability_rate', 15)

            kpis['profitability'] = {
                'total_revenue': float(total_revenue),
                'total_cost': float(total_cost),
                'land_cost': float(land_cost),
                'profit': float(profit),
                'roi': roi,
                'profit_margin': profit_margin,
                'developer_rate': developer_rate,
                'is_profitable': profit > 0
            }

            # Area Metrics - Calculate from revenue_forecast and cost_forecast
            revenue_forecast = data_inputs.revenue_forecast or {}
            cost_forecast = data_inputs.cost_forecast or {}
            revenue_residential = revenue_forecast.get('revenue_residential', [])
            revenue_commercial = revenue_forecast.get('revenue_commercial', [])
            cost_data = cost_forecast.get('cost_forecast_data', [])

            # Calculate total areas and values from residential units
            total_main_area = Decimal('0')  # שטח פלדלת
            total_balcony_area = Decimal('0')
            total_roof_terrace_area = Decimal('0')
            total_value_no_vat = Decimal('0')
            total_value_with_vat = Decimal('0')
            unit_count = 0
            avg_price_per_sqm_equiv = Decimal('0')
            price_per_sqm_equiv_sum = Decimal('0')

            for unit in revenue_residential:
                try:
                    area_main = Decimal(str(unit.get('area_main', 0) or 0))
                    balcony = Decimal(str(unit.get('balcony_sun', 0) or 0))
                    roof_terrace = Decimal(str(unit.get('roof_terrace', 0) or 0))
                    value_no_vat = Decimal(str(unit.get('total_value_no_vat', 0) or 0))
                    value_with_vat = Decimal(str(unit.get('total_value_with_vat', 0) or 0))
                    price_sqm_equiv = Decimal(str(unit.get('price_per_sqm_equiv', 0) or 0))

                    total_main_area += area_main
                    total_balcony_area += balcony
                    total_roof_terrace_area += roof_terrace
                    total_value_no_vat += value_no_vat
                    total_value_with_vat += value_with_vat
                    price_per_sqm_equiv_sum += price_sqm_equiv
                    unit_count += 1
                except:
                    pass

            # Add commercial areas
            total_commercial_area = Decimal('0')
            total_commercial_value = Decimal('0')
            for unit in revenue_commercial:
                try:
                    area_gross = Decimal(str(unit.get('area_gross', 0) or 0))
                    value_no_vat = Decimal(str(unit.get('total_value_no_vat', 0) or 0))
                    total_commercial_area += area_gross
                    total_commercial_value += value_no_vat
                except:
                    pass

            # Calculate cost breakdowns
            construction_cost = Decimal('0')
            land_acquisition_cost = Decimal('0')
            financing_cost = Decimal('0')
            marketing_cost = Decimal('0')
            professional_fees = Decimal('0')
            other_costs = Decimal('0')

            for cost_item in cost_data:
                try:
                    cost = Decimal(str(cost_item.get('cost_no_vat', 0) or 0))
                    category = str(cost_item.get('category', '')).lower()
                    item_name = str(cost_item.get('item', '')).lower()

                    if 'בנייה' in category or 'construction' in category or 'קבלן' in category:
                        construction_cost += cost
                    elif 'קרקע' in category or 'land' in category:
                        land_acquisition_cost += cost
                    elif 'מימון' in category or 'financing' in category or 'ריבית' in category:
                        financing_cost += cost
                    elif 'שיווק' in category or 'marketing' in category or 'מכירות' in category:
                        marketing_cost += cost
                    elif 'יועצים' in category or 'professional' in category or 'אדריכל' in category or 'מהנדס' in category:
                        professional_fees += cost
                    else:
                        other_costs += cost
                except:
                    pass

            # Calculate equivalent area (main + 50% balcony + 30% roof terrace)
            total_equivalent_area = total_main_area + (total_balcony_area * Decimal('0.5')) + (total_roof_terrace_area * Decimal('0.3'))
            total_building_area = total_main_area + total_balcony_area + total_roof_terrace_area + total_commercial_area

            # Calculate per-sqm metrics
            avg_unit_value = float(total_value_with_vat) / unit_count if unit_count > 0 else 0
            avg_unit_area = float(total_main_area) / unit_count if unit_count > 0 else 0
            avg_price_sqm_equiv = float(price_per_sqm_equiv_sum) / unit_count if unit_count > 0 else 0

            revenue_per_sqm = float(total_value_no_vat) / float(total_main_area) if total_main_area > 0 else 0
            cost_per_sqm = float(total_cost) / float(total_building_area) if total_building_area > 0 else 0
            profit_per_sqm = float(profit) / float(total_main_area) if total_main_area > 0 else 0
            construction_cost_per_sqm = float(construction_cost) / float(total_building_area) if total_building_area > 0 else 0
            land_cost_per_sqm = float(land_cost) / float(total_building_area) if total_building_area > 0 else 0

            # Breakeven calculation: what % of units need to be sold to cover costs
            breakeven_sales_percent = round((float(total_cost) / float(total_value_no_vat) * 100), 1) if total_value_no_vat > 0 else 0

            kpis['area_metrics'] = {
                # Area summaries
                'total_main_area': float(total_main_area),
                'total_balcony_area': float(total_balcony_area),
                'total_roof_terrace_area': float(total_roof_terrace_area),
                'total_commercial_area': float(total_commercial_area),
                'total_equivalent_area': float(total_equivalent_area),
                'total_building_area': float(total_building_area),
                'unit_count': unit_count,

                # Per-unit metrics
                'avg_unit_value': round(avg_unit_value, 0),
                'avg_unit_area': round(avg_unit_area, 1),

                # Per-sqm metrics (key profitability indicators)
                'avg_price_per_sqm_equiv': round(avg_price_sqm_equiv, 0),  # שווי למ"ר אקוו' כולל מע"מ
                'revenue_per_sqm': round(revenue_per_sqm, 0),  # הכנסה למ"ר
                'cost_per_sqm': round(cost_per_sqm, 0),  # עלות למ"ר
                'profit_per_sqm': round(profit_per_sqm, 0),  # רווח למ"ר
                'construction_cost_per_sqm': round(construction_cost_per_sqm, 0),  # עלות בנייה למ"ר
                'land_cost_per_sqm': round(land_cost_per_sqm, 0),  # עלות קרקע למ"ר

                # Cost breakdown
                'cost_breakdown': {
                    'construction': float(construction_cost),
                    'land': float(land_acquisition_cost),
                    'financing': float(financing_cost),
                    'marketing': float(marketing_cost),
                    'professional_fees': float(professional_fees),
                    'other': float(other_costs)
                },

                # Breakeven analysis
                'breakeven_sales_percent': breakeven_sales_percent,
                'is_above_breakeven': kpis.get('sales', {}).get('sales_percent', 0) >= breakeven_sales_percent if breakeven_sales_percent > 0 else False
            }

            # Timeline
            dates = data_inputs.dates or {}
            construction_start = dates.get('construction_start_date')
            duration_months = int(dates.get('construction_duration_months', 36) or 36)
            permit_date = dates.get('construction_permit_date')

            if construction_start:
                start_date = datetime.strptime(construction_start, '%Y-%m-%d').date()
                end_date = start_date + relativedelta(months=duration_months)
                today = date.today()

                if today < start_date:
                    status_text = 'טרם התחיל'
                    days_remaining = (end_date - start_date).days
                    elapsed_months = 0
                    progress_percent = 0
                elif today > end_date:
                    status_text = 'הושלם'
                    days_remaining = 0
                    elapsed_months = duration_months
                    progress_percent = 100
                else:
                    status_text = 'בבנייה'
                    days_remaining = (end_date - today).days
                    total_days = (end_date - start_date).days
                    elapsed_days = (today - start_date).days
                    progress_percent = round((elapsed_days / total_days) * 100, 1) if total_days > 0 else 0
                    diff = relativedelta(today, start_date)
                    elapsed_months = diff.years * 12 + diff.months

                kpis['timeline'] = {
                    'start_date': construction_start,
                    'end_date': end_date.strftime('%Y-%m-%d'),
                    'permit_date': permit_date,
                    'duration_months': duration_months,
                    'elapsed_months': elapsed_months,
                    'remaining_months': max(0, duration_months - elapsed_months),
                    'days_remaining': days_remaining,
                    'progress_percent': progress_percent,
                    'status': status_text,
                    'on_schedule': progress_percent <= (elapsed_months / duration_months * 100) + 10 if duration_months > 0 else True
                }
            else:
                kpis['timeline'] = {
                    'status': 'לא הוגדר',
                    'message': 'תאריך התחלת בנייה לא הוגדר'
                }

        except ProjectDataInputs.DoesNotExist:
            kpis['profitability'] = {'message': 'אין נתוני פרויקט'}
            kpis['timeline'] = {'message': 'אין נתוני פרויקט'}

        # Sales data - try ApartmentInventory first, then fallback to Section 7 data
        try:
            apartments = ApartmentInventory.objects.filter(project=project)
            total_apartments = apartments.count()

            # If ApartmentInventory is empty, try to use Section 7 (revenue_forecast) data
            if total_apartments == 0:
                try:
                    data_inputs = project.data_inputs
                    revenue_forecast = data_inputs.revenue_forecast or {}
                    revenue_residential = revenue_forecast.get('revenue_residential', [])

                    if revenue_residential and len(revenue_residential) > 0:
                        # Count by status from Section 7 data
                        total_apartments = len(revenue_residential)
                        sold_apartments = 0
                        available_apartments = 0
                        reserved_apartments = 0
                        total_sales_value_calc = Decimal('0')

                        for unit in revenue_residential:
                            status = str(unit.get('status', '')).lower()
                            unit_value = Decimal(str(unit.get('total_value_with_vat', 0) or 0))
                            total_sales_value_calc += unit_value

                            if 'נמכר' in status or 'sold' in status.lower():
                                sold_apartments += 1
                            elif 'שמור' in status or 'משוריין' in status or 'reserved' in status.lower():
                                reserved_apartments += 1
                            else:
                                # לשיווק, להשכרה, בעלים, תמורה all count as available
                                available_apartments += 1

                        avg_price = float(total_sales_value_calc) / total_apartments if total_apartments > 0 else 0

                        kpis['sales'] = {
                            'total_apartments': total_apartments,
                            'sold_apartments': sold_apartments,
                            'available_apartments': available_apartments,
                            'reserved_apartments': reserved_apartments,
                            'sales_percent': round((sold_apartments / total_apartments * 100), 1) if total_apartments > 0 else 0,
                            'total_sales_value': float(total_sales_value_calc),
                            'average_price': round(avg_price, 0),
                            'total_scheduled_payments': 0,
                            'total_collected': 0,
                            'collection_rate': 0,
                            'outstanding_balance': 0,
                            'source': 'section7'  # Indicate data source
                        }
                except Exception:
                    pass

            # If we already have kpis['sales'] from Section 7, skip ApartmentInventory
            if 'sales' not in kpis or kpis.get('sales', {}).get('total_apartments', 0) == 0:
                sold_apartments = apartments.filter(unit_status='SOLD').count()
                available_apartments = apartments.filter(unit_status='FOR_SALE').count()
                reserved_apartments = apartments.filter(unit_status='RESERVED').count()

                # Sales transactions
                sales = SalesTransaction.objects.filter(apartment__project=project)
                total_sales_value = sales.aggregate(total=Sum('final_price'))['total'] or Decimal('0')

                # Payment collection
                payments = CustomerPaymentSchedule.objects.filter(sales_transaction__apartment__project=project)
                total_scheduled = payments.aggregate(total=Sum('scheduled_amount'))['total'] or Decimal('0')
                total_paid = payments.filter(payment_status='PAID').aggregate(total=Sum('actual_amount'))['total'] or Decimal('0')

                collection_rate = round((float(total_paid) / float(total_scheduled) * 100), 1) if total_scheduled > 0 else 0

                # Average price per apartment
                avg_price = float(total_sales_value) / sold_apartments if sold_apartments > 0 else 0

                kpis['sales'] = {
                    'total_apartments': total_apartments,
                    'sold_apartments': sold_apartments,
                    'available_apartments': available_apartments,
                    'reserved_apartments': reserved_apartments,
                    'sales_percent': round((sold_apartments / total_apartments * 100), 1) if total_apartments > 0 else 0,
                    'total_sales_value': float(total_sales_value),
                    'average_price': round(avg_price, 0),
                    'total_scheduled_payments': float(total_scheduled),
                    'total_collected': float(total_paid),
                    'collection_rate': collection_rate,
                    'outstanding_balance': float(total_scheduled - total_paid)
                }
        except Exception as e:
            kpis['sales'] = {'error': str(e)}

        # Construction progress
        try:
            construction = ConstructionProgress.objects.get(project=project)
            kpis['construction'] = {
                'overall_progress': float(construction.overall_completion_percentage or 0),
                'total_contract': float(construction.total_contract_amount or 0),
                'total_spent': float(construction.total_spent_to_date or 0),
                'spend_rate': round((float(construction.total_spent_to_date or 0) / float(construction.total_contract_amount) * 100), 1) if construction.total_contract_amount else 0
            }
        except ConstructionProgress.DoesNotExist:
            kpis['construction'] = {'overall_progress': 0, 'message': 'אין נתוני התקדמות בנייה'}

        # Equity and cashflow
        try:
            total_equity = EquityDeposit.objects.filter(project=project).aggregate(
                total=Sum('amount')
            )['total'] or Decimal('0')

            # Bank transactions summary
            bank_txns = BankTransaction.objects.filter(project=project, status='APPROVED')
            total_income = bank_txns.filter(transaction_type='CREDIT').aggregate(
                total=Sum('amount')
            )['total'] or Decimal('0')
            total_expenses = bank_txns.filter(transaction_type='DEBIT').aggregate(
                total=Sum('amount')
            )['total'] or Decimal('0')

            kpis['cashflow'] = {
                'total_equity': float(total_equity),
                'bank_income': float(total_income),
                'bank_expenses': float(total_expenses),
                'net_cashflow': float(total_income - total_expenses),
                'cash_balance': float(total_income - total_expenses + total_equity)
            }
        except Exception as e:
            kpis['cashflow'] = {'error': str(e)}

        return Response(kpis)


class MonthlyMonitoringView(APIView):
    """
    API view for monthly monitoring data - aggregates all key metrics for a specific month.
    """
    def get(self, request, project_pk=None, *args, **kwargs):
        from apps.sales.models import SalesTransaction, ApartmentInventory, CustomerPaymentSchedule
        from django.db.models import Sum, Count
        from dateutil.relativedelta import relativedelta
        from calendar import monthrange

        # Get month parameter (format: YYYY-MM)
        month_str = request.query_params.get('month')
        if month_str:
            try:
                year, month = map(int, month_str.split('-'))
                report_date = datetime(year, month, 1).date()
            except (ValueError, TypeError):
                report_date = datetime.now().date().replace(day=1)
        else:
            report_date = datetime.now().date().replace(day=1)

        # Calculate month boundaries
        month_start = report_date
        month_end = report_date + relativedelta(months=1) - relativedelta(days=1)

        try:
            project = Project.objects.get(pk=project_pk, is_active=True)
        except Project.DoesNotExist:
            return Response(
                {'error': 'Project not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        monitoring_data = {
            'project_id': project.id,
            'project_name': project.project_name,
            'project_code': project.project_id,
            'month': month_str or report_date.strftime('%Y-%m'),
            'month_display': report_date.strftime('%B %Y'),
            'generated_at': datetime.now().isoformat()
        }

        # ============================================
        # Financial Overview
        # ============================================
        try:
            data_inputs = project.data_inputs
            land_value = data_inputs.land_value or {}
            cost_forecast = data_inputs.cost_forecast or {}

            total_budget = Decimal(str(land_value.get('total_cost_project', 0) or 0))
            total_revenue = Decimal(str(land_value.get('total_income_combined_no_vat', 0) or 0))

            # Get monthly expenses from bank transactions
            monthly_expenses = BankTransaction.objects.filter(
                project=project,
                transaction_type='DEBIT',
                status='APPROVED',
                transaction_date__gte=month_start,
                transaction_date__lte=month_end
            ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

            # Get total expenses to date
            total_expenses = BankTransaction.objects.filter(
                project=project,
                transaction_type='DEBIT',
                status='APPROVED'
            ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

            # Calculate planned monthly (simple division for now)
            construction_months = 36  # default
            try:
                dates = data_inputs.dates or {}
                construction_months = int(dates.get('construction_duration_months', 36) or 36)
            except:
                pass

            planned_monthly = float(total_budget) / construction_months if construction_months > 0 else 0
            variance = float(monthly_expenses) - planned_monthly
            variance_percent = (variance / planned_monthly * 100) if planned_monthly > 0 else 0

            monitoring_data['financial'] = {
                'total_budget': float(total_budget),
                'total_revenue': float(total_revenue),
                'spent_to_date': float(total_expenses),
                'remaining_budget': float(total_budget - total_expenses),
                'monthly_expenses': float(monthly_expenses),
                'planned_monthly': round(planned_monthly, 0),
                'variance': round(variance, 0),
                'variance_percent': round(variance_percent, 1),
                'budget_utilization': round((float(total_expenses) / float(total_budget) * 100), 1) if total_budget > 0 else 0
            }
        except Exception as e:
            monitoring_data['financial'] = {'error': str(e)}

        # ============================================
        # Construction Progress
        # ============================================
        try:
            construction = ConstructionProgress.objects.get(project=project)
            overall_progress = float(construction.overall_completion_percentage or 0)

            # Get progress snapshot for comparison (previous month)
            prev_month = month_start - relativedelta(months=1)
            prev_snapshot = ConstructionProgressSnapshot.objects.filter(
                project=project,
                year=prev_month.year,
                month=prev_month.month
            ).first()

            prev_progress = float(prev_snapshot.overall_completion if prev_snapshot else 0)
            monthly_progress = overall_progress - prev_progress

            # Get milestones from progress_details
            milestones = []
            progress_details = construction.progress_details or []
            for detail in progress_details[:10]:  # Limit to 10 milestones
                planned = float(detail.get('planned_percentage', 0) or 0)
                actual = float(detail.get('actual_percentage', 0) or 0)

                if actual >= planned:
                    status_val = 'completed' if actual >= 100 else ('ahead' if actual > planned else 'in_progress')
                else:
                    status_val = 'delayed' if actual < planned - 5 else 'in_progress'

                milestones.append({
                    'name': detail.get('item', 'Unknown'),
                    'planned': planned,
                    'actual': actual,
                    'status': status_val
                })

            monitoring_data['construction'] = {
                'overall_progress': overall_progress,
                'monthly_actual': round(monthly_progress, 1),
                'monthly_target': 3.0,  # Default monthly target
                'variance': round(monthly_progress - 3.0, 1),
                'total_contract': float(construction.total_contract_amount or 0),
                'total_spent': float(construction.total_spent_to_date or 0),
                'milestones': milestones
            }
        except ConstructionProgress.DoesNotExist:
            monitoring_data['construction'] = {
                'overall_progress': 0,
                'monthly_actual': 0,
                'monthly_target': 0,
                'variance': 0,
                'milestones': [],
                'message': 'No construction progress data'
            }

        # ============================================
        # Bank/Equity Status
        # ============================================
        try:
            # Equity collected
            equity_collected = EquityDeposit.objects.filter(
                project=project
            ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

            # Get equity target from financing data
            try:
                financing = data_inputs.financing or {}
                equity_target = Decimal(str(financing.get('required_equity', 0) or 0))
            except:
                equity_target = Decimal('0')

            # Bank releases (credit transactions)
            bank_releases = BankTransaction.objects.filter(
                project=project,
                transaction_type='CREDIT',
                status='APPROVED'
            ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

            # Pending releases
            pending_releases = BankTransaction.objects.filter(
                project=project,
                transaction_type='CREDIT',
                status='PENDING'
            ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

            monitoring_data['bank'] = {
                'equity_collected': float(equity_collected),
                'equity_target': float(equity_target) if equity_target > 0 else float(equity_collected) * 1.2,
                'equity_percent': round((float(equity_collected) / float(equity_target) * 100), 1) if equity_target > 0 else 0,
                'bank_releases': float(bank_releases),
                'pending_releases': float(pending_releases),
                'next_release_date': None  # Would come from bank schedule if available
            }
        except Exception as e:
            monitoring_data['bank'] = {'error': str(e)}

        # ============================================
        # Sales Status
        # ============================================
        try:
            apartments = ApartmentInventory.objects.filter(project=project)
            total_units = apartments.count()
            sold_units = apartments.filter(unit_status='SOLD').count()

            # If no ApartmentInventory, try Section 7 data
            if total_units == 0:
                try:
                    revenue_forecast = data_inputs.revenue_forecast or {}
                    revenue_residential = revenue_forecast.get('revenue_residential', [])
                    total_units = len(revenue_residential)
                    sold_units = sum(1 for u in revenue_residential if 'נמכר' in str(u.get('status', '')).lower() or 'sold' in str(u.get('status', '')).lower())
                except:
                    pass

            # Monthly sales (transactions in this month)
            monthly_sales = SalesTransaction.objects.filter(
                apartment__project=project,
                sale_date__gte=month_start,
                sale_date__lte=month_end
            ).count()

            monitoring_data['sales'] = {
                'total_units': total_units,
                'sold_units': sold_units,
                'available_units': total_units - sold_units,
                'sales_percent': round((sold_units / total_units * 100), 1) if total_units > 0 else 0,
                'monthly_sales': monthly_sales,
                'target_sales': 2  # Default monthly target
            }
        except Exception as e:
            monitoring_data['sales'] = {'error': str(e)}

        # ============================================
        # Alerts and Issues
        # ============================================
        alerts = []

        # Check construction delays
        if monitoring_data.get('construction', {}).get('variance', 0) < -2:
            alerts.append({
                'type': 'warning',
                'message': 'עיכוב בהתקדמות הבנייה מתחת ליעד החודשי',
                'priority': 'high'
            })

        # Check budget overruns
        if monitoring_data.get('financial', {}).get('variance_percent', 0) > 10:
            alerts.append({
                'type': 'warning',
                'message': 'חריגה בהוצאות מעל 10% מהתכנון',
                'priority': 'high'
            })

        # Check equity collection
        if monitoring_data.get('bank', {}).get('equity_percent', 100) < 80:
            alerts.append({
                'type': 'info',
                'message': 'גביית הון עצמי מתחת ל-80% מהיעד',
                'priority': 'medium'
            })

        # Check sales progress
        if monitoring_data.get('sales', {}).get('sales_percent', 0) < 50:
            alerts.append({
                'type': 'info',
                'message': 'מכירות מתחת ל-50%',
                'priority': 'medium'
            })

        # Positive alerts
        if monitoring_data.get('construction', {}).get('variance', 0) > 0:
            alerts.append({
                'type': 'success',
                'message': 'התקדמות הבנייה מעל היעד החודשי',
                'priority': 'low'
            })

        monitoring_data['alerts'] = alerts

        # ============================================
        # Notes (stored separately per month)
        # ============================================
        monitoring_data['notes'] = ''  # Would be stored in a MonthlyReport model

        return Response(monitoring_data)


class BankChoicesView(APIView):
    """
    An API view to provide the list of bank choices.
    """
    def get(self, request, *args, **kwargs):
        formatted_banks = [
            {"value": choice[0], "label": choice[1]}
            for choice in Project.BANK_CHOICES
        ]
        return Response(formatted_banks)


class TransactionCategoryChoicesView(APIView):
    """
    An API view to provide the list of transaction category choices.
    """
    def get(self, request, *args, **kwargs):
        formatted_categories = [
            {"value": choice[0], "label": choice[1]}
            for choice in BankTransaction.CATEGORY_CHOICES
        ]
        return Response(formatted_categories)


class BankTransactionViewSet(viewsets.ModelViewSet):
    """ViewSet for managing bank transactions"""
    queryset = BankTransaction.objects.all()
    serializer_class = BankTransactionSerializer
    parser_classes = [JSONParser, MultiPartParser, FormParser]

    def get_queryset(self):
        queryset = BankTransaction.objects.all()

        # Filter by project if specified
        project_id = self.kwargs.get('project_pk') or self.request.query_params.get('project')
        if project_id:
            queryset = queryset.filter(project_id=project_id)

        # Filter by status if specified
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        return queryset.order_by('-transaction_date', '-id')

    @action(detail=False, methods=['get'], url_path='project/(?P<project_pk>[^/.]+)/transactions')
    def project_transactions(self, request, project_pk=None):
        """Get all transactions for a specific project"""
        queryset = BankTransaction.objects.filter(project_id=project_pk)

        # Filter by status if specified
        status_filter = request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        queryset = queryset.order_by('-transaction_date', '-id')
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='project/(?P<project_pk>[^/.]+)/summary')
    def project_summary(self, request, project_pk=None):
        """Get summary statistics for a project's bank transactions"""
        queryset = BankTransaction.objects.filter(project_id=project_pk)

        total_transactions = queryset.count()
        pending_count = queryset.filter(status='PENDING').count()
        approved_count = queryset.filter(status='APPROVED').count()
        rejected_count = queryset.filter(status='REJECTED').count()

        total_debit = queryset.filter(transaction_type='DEBIT').aggregate(
            total=Sum('amount')
        )['total'] or Decimal('0')

        total_credit = queryset.filter(transaction_type='CREDIT').aggregate(
            total=Sum('amount')
        )['total'] or Decimal('0')

        return Response({
            'total_transactions': total_transactions,
            'pending_count': pending_count,
            'approved_count': approved_count,
            'rejected_count': rejected_count,
            'total_debit': float(total_debit),
            'total_credit': float(total_credit),
            'net_balance': float(total_credit - total_debit)
        })

    @action(detail=False, methods=['post'], url_path='project/(?P<project_pk>[^/.]+)/upload')
    def upload_statement(self, request, project_pk=None):
        """Upload a bank statement Excel file"""
        serializer = BankStatementUploadSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        file = serializer.validated_data['file']

        try:
            project = Project.objects.get(pk=project_pk)
        except Project.DoesNotExist:
            return Response(
                {'error': 'Project not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        try:
            # Parse the Excel file
            transactions = self._parse_bank_statement(file, project)

            # Save transactions
            created_count = 0
            bank_name = None
            for txn_data in transactions:
                txn_data['project'] = project
                BankTransaction.objects.create(**txn_data)
                created_count += 1
                if not bank_name:
                    bank_name = txn_data.get('bank', 'Unknown')

            return Response({
                'message': f'Successfully imported {created_count} transactions',
                'transaction_count': created_count,
                'bank': bank_name,
                'file_name': file.name
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response(
                {'error': f'Failed to parse file: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

    def _parse_bank_statement(self, file, project):
        """Parse bank statement Excel file and extract transactions"""
        workbook = openpyxl.load_workbook(file, data_only=True)
        sheet = workbook.active

        transactions = []
        bank = self._detect_bank(sheet)

        # Find header row and parse based on bank format
        # This is a simplified parser - in production, each bank would have specific parsing logic
        header_row = None
        for row_num, row in enumerate(sheet.iter_rows(min_row=1, max_row=20), start=1):
            cell_values = [str(cell.value).lower() if cell.value else '' for cell in row]
            if any('תאריך' in val or 'date' in val for val in cell_values):
                header_row = row_num
                break

        if not header_row:
            header_row = 1  # Default to first row

        # Parse data rows
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or not any(row):
                continue

            # Try to extract transaction data
            try:
                txn = self._extract_transaction_from_row(row, bank)
                if txn:
                    txn['uploaded_file'] = file.name
                    transactions.append(txn)
            except Exception:
                continue

        return transactions

    def _detect_bank(self, sheet):
        """Detect which bank the statement is from"""
        # Check first few rows for bank identifiers
        for row in sheet.iter_rows(min_row=1, max_row=10, values_only=True):
            row_text = ' '.join(str(cell) if cell else '' for cell in row).lower()

            if 'לאומי' in row_text or 'leumi' in row_text:
                return 'LEUMI'
            elif 'הפועלים' in row_text or 'hapoalim' in row_text:
                return 'HAPOALIM'
            elif 'דיסקונט' in row_text or 'discount' in row_text:
                return 'DISCOUNT'
            elif 'מזרחי' in row_text or 'mizrahi' in row_text:
                return 'MIZRAHI'

        return 'OTHER'

    def _extract_transaction_from_row(self, row, bank):
        """Extract transaction data from a row"""
        # Basic extraction - assumes common format
        # In production, this would be bank-specific

        date_val = None
        description = None
        amount = None
        balance = None

        for i, cell in enumerate(row):
            if cell is None:
                continue

            # Try to detect date
            if isinstance(cell, datetime):
                date_val = cell.date()
            elif isinstance(cell, str) and not date_val:
                # Try parsing date string
                try:
                    for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d.%m.%Y']:
                        try:
                            date_val = datetime.strptime(cell, fmt).date()
                            break
                        except ValueError:
                            continue
                except:
                    pass

            # Try to detect numeric values (amount/balance)
            if isinstance(cell, (int, float)) and cell != 0:
                if amount is None:
                    amount = Decimal(str(cell))
                elif balance is None:
                    balance = Decimal(str(cell))

            # Try to detect description (string that's not a date)
            if isinstance(cell, str) and len(cell) > 3 and not date_val:
                if description is None or len(cell) > len(description):
                    description = cell

        if not date_val or amount is None:
            return None

        # Determine transaction type based on amount sign
        if amount < 0:
            transaction_type = 'DEBIT'
            amount = abs(amount)
        else:
            transaction_type = 'CREDIT'

        return {
            'bank': bank,
            'transaction_date': date_val,
            'description': description or 'No description',
            'amount': amount,
            'balance': balance,
            'transaction_type': transaction_type,
            'status': 'PENDING'
        }

    @action(detail=True, methods=['put'], url_path='approve')
    def approve_transaction(self, request, pk=None):
        """Approve or reject a single transaction"""
        transaction = self.get_object()
        serializer = BankTransactionApprovalSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        transaction.status = serializer.validated_data['status']
        transaction.approval_notes = serializer.validated_data.get('approval_notes', '')
        transaction.approved_by = serializer.validated_data.get('approved_by', '')
        transaction.approved_date = timezone.now()

        if 'category' in serializer.validated_data:
            transaction.category = serializer.validated_data['category']
        if 'is_construction_related' in serializer.validated_data:
            transaction.is_construction_related = serializer.validated_data['is_construction_related']

        transaction.save()

        return Response(BankTransactionSerializer(transaction).data)

    @action(detail=False, methods=['post'], url_path='bulk-approve')
    def bulk_approve(self, request):
        """Approve or reject multiple transactions"""
        serializer = BankTransactionBulkApprovalSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        transaction_ids = serializer.validated_data['transaction_ids']
        new_status = serializer.validated_data['status']
        approval_notes = serializer.validated_data.get('approval_notes', '')
        approved_by = serializer.validated_data.get('approved_by', '')

        updated_count = BankTransaction.objects.filter(
            id__in=transaction_ids
        ).update(
            status=new_status,
            approval_notes=approval_notes,
            approved_by=approved_by,
            approved_date=timezone.now()
        )

        return Response({
            'updated_count': updated_count,
            'status': new_status
        })


class ConstructionProgressViewSet(viewsets.ModelViewSet):
    """ViewSet for managing construction progress"""
    queryset = ConstructionProgress.objects.all()
    serializer_class = ConstructionProgressSerializer

    @action(detail=False, methods=['get'], url_path='project/(?P<project_pk>[^/.]+)/progress')
    def project_progress(self, request, project_pk=None):
        """Get construction progress for a specific project"""
        try:
            progress = ConstructionProgress.objects.get(project_id=project_pk)
            serializer = self.get_serializer(progress)
            return Response(serializer.data)
        except ConstructionProgress.DoesNotExist:
            return Response(
                {'detail': 'No construction progress found for this project'},
                status=status.HTTP_404_NOT_FOUND
            )

    @action(detail=False, methods=['get'], url_path='project/(?P<project_pk>[^/.]+)/snapshots')
    def project_snapshots(self, request, project_pk=None):
        """Get all progress snapshots for a project"""
        snapshots = ConstructionProgressSnapshot.objects.filter(
            project_id=project_pk
        ).order_by('-year', '-month')
        serializer = ConstructionProgressSnapshotSerializer(snapshots, many=True)
        return Response(serializer.data)


class EquityDepositViewSet(viewsets.ModelViewSet):
    """ViewSet for managing equity deposits"""
    queryset = EquityDeposit.objects.all()
    serializer_class = EquityDepositSerializer

    def get_queryset(self):
        queryset = EquityDeposit.objects.all()

        # Filter by project if specified
        project_id = self.kwargs.get('project_pk') or self.request.query_params.get('project')
        if project_id:
            queryset = queryset.filter(project_id=project_id)

        return queryset.order_by('-deposit_date', '-id')

    @action(detail=False, methods=['get'], url_path='project/(?P<project_pk>[^/.]+)/deposits')
    def project_deposits(self, request, project_pk=None):
        """Get all equity deposits for a specific project"""
        deposits = EquityDeposit.objects.filter(project_id=project_pk).order_by('-deposit_date', '-id')
        serializer = self.get_serializer(deposits, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='project/(?P<project_pk>[^/.]+)/summary')
    def project_summary(self, request, project_pk=None):
        """Get summary of equity deposits for a project"""
        deposits = EquityDeposit.objects.filter(project_id=project_pk)

        total_deposited = deposits.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        deposit_count = deposits.count()

        # Get deposits by source
        by_source = deposits.values('source').annotate(
            total=Sum('amount'),
            count=Count('id')
        )

        return Response({
            'total_deposited': float(total_deposited),
            'deposit_count': deposit_count,
            'by_source': list(by_source)
        })


from django.shortcuts import get_object_or_404


class ProjectDataInputsViewSet(viewsets.ViewSet):
    """ViewSet for managing project data inputs"""

    # Field name mapping from section IDs to model field names
    FIELD_MAPPING = {
        'property-details': 'property_details',
        'developer': 'developer',
        'dates': 'dates',
        'planned-area': 'property_details',  # stored in property_details
        'financing': 'financing',
        'fixed-rates': 'fixed_rates',
        'revenue-forecast': 'revenue_forecast',
        'cost-forecast': 'cost_forecast',
        'construction-classification': 'construction_classification',
        'insurance': 'insurance',
        'guarantees': 'guarantees',
        'profitability': 'profitability',
        'land-value': 'land_value',
        'sensitivity-analysis': 'sensitivity_analysis',
        'monthly-cashflow': 'monthly_cashflow',
    }

    def _get_field_name(self, section_id):
        """Convert section ID to model field name"""
        return self.FIELD_MAPPING.get(section_id, section_id.replace('-', '_'))

    @action(detail=False, methods=['get', 'post'], url_path='project/(?P<project_id>[^/.]+)/(?P<section_id>[^/.]+)')
    def handle_section(self, request, project_id=None, section_id=None):
        """Generic handler for all data input sections"""
        project = get_object_or_404(Project, id=project_id)
        data_inputs, created = ProjectDataInputs.objects.get_or_create(project=project)

        field_name = self._get_field_name(section_id)

        if request.method == 'GET':
            # Return the data for this section
            try:
                data = getattr(data_inputs, field_name, None)
                return Response({'data': data})
            except AttributeError:
                return Response({'data': None})
        else:
            # Save the data for this section
            try:
                new_data = request.data.get('data', request.data)
                setattr(data_inputs, field_name, new_data)
                data_inputs.save()
                return Response({'status': 'saved'}, status=status.HTTP_200_OK)
            except Exception as e:
                return Response(
                    {'error': str(e)},
                    status=status.HTTP_400_BAD_REQUEST
                )


class ProjectDocumentViewSet(viewsets.ModelViewSet):
    """ViewSet for managing project documents"""
    queryset = ProjectDocument.objects.all()
    serializer_class = ProjectDocumentSerializer
    parser_classes = [MultiPartParser, FormParser]

    def get_queryset(self):
        queryset = ProjectDocument.objects.all()

        # Filter by project if specified
        project_id = self.kwargs.get('project_pk') or self.request.query_params.get('project')
        if project_id:
            queryset = queryset.filter(project_id=project_id)

        # Filter by category if specified
        category_filter = self.request.query_params.get('category')
        if category_filter:
            queryset = queryset.filter(category=category_filter)

        return queryset.order_by('-created_at')

    @action(detail=False, methods=['get'], url_path='project/(?P<project_pk>[^/.]+)/documents')
    def project_documents(self, request, project_pk=None):
        """Get all documents for a specific project"""
        queryset = ProjectDocument.objects.filter(project_id=project_pk)

        # Filter by category if specified
        category_filter = request.query_params.get('category')
        if category_filter:
            queryset = queryset.filter(category=category_filter)

        queryset = queryset.order_by('-created_at')
        serializer = self.get_serializer(queryset, many=True, context={'request': request})
        return Response(serializer.data)

    @action(detail=False, methods=['post'], url_path='project/(?P<project_pk>[^/.]+)/upload')
    def upload_document(self, request, project_pk=None):
        """Upload a document for a project"""
        try:
            project = Project.objects.get(pk=project_pk)
        except Project.DoesNotExist:
            return Response(
                {'error': 'Project not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        serializer = ProjectDocumentUploadSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        uploaded_file = serializer.validated_data['file']

        # Create the document
        document = ProjectDocument.objects.create(
            project=project,
            name=serializer.validated_data['name'],
            description=serializer.validated_data.get('description', ''),
            category=serializer.validated_data.get('category', 'OTHER'),
            file=uploaded_file,
            file_name=uploaded_file.name,
            file_size=uploaded_file.size,
            file_type=uploaded_file.content_type or 'application/octet-stream',
            document_date=serializer.validated_data.get('document_date'),
            tags=serializer.validated_data.get('tags', []),
            uploaded_by=request.user.get_full_name() if request.user.is_authenticated else 'Anonymous'
        )

        return Response(
            ProjectDocumentSerializer(document, context={'request': request}).data,
            status=status.HTTP_201_CREATED
        )

    @action(detail=False, methods=['get'], url_path='project/(?P<project_pk>[^/.]+)/summary')
    def project_summary(self, request, project_pk=None):
        """Get summary of documents for a project"""
        documents = ProjectDocument.objects.filter(project_id=project_pk)

        total_count = documents.count()
        total_size = documents.aggregate(total=Sum('file_size'))['total'] or 0

        # Get counts by category
        by_category = documents.values('category').annotate(
            count=Count('id')
        )

        return Response({
            'total_count': total_count,
            'total_size': total_size,
            'total_size_mb': round(total_size / (1024 * 1024), 2),
            'by_category': list(by_category)
        })

    @action(detail=False, methods=['get'])
    def categories(self, request):
        """Get list of document categories"""
        categories = [
            {'value': choice[0], 'label': choice[1]}
            for choice in ProjectDocument.CATEGORY_CHOICES
        ]
        return Response(categories)