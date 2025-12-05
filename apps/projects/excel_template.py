from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Define all sections with their headers
TEMPLATE_SECTIONS = {
    'property_details': {
        'sheet_name': 'פרטי נכס',
        'fields': {
            'גוש': 'gush',
            'מספר תכנית': 'plan_number',
            'היזם': 'developer',
            'זכויות': 'rights'
        },
        'table_headers': ['חלקה', 'שטח רשום במ"ר', 'תא שטח', 'ייעוד']
    },
    'project_description': {
        'sheet_name': 'תיאור פרויקט',
        'fields': {
            'תיאור טקסטואלי': 'description_text'
        },
        'table_headers': ['פריט', 'ערך']
    },
    'timeline': {
        'sheet_name': 'לוח זמנים',
        'fields': {
            'קבלת היתר בנייה': 'permit_date',
            'התחלת ביצוע': 'construction_start',
            'משך הביצוע (חודשים)': 'duration_months'
        }
    },
    'sales_timeline': {
        'sheet_name': 'לוח מכירות',
        'table_headers': ['אבן דרך', 'תאריך משוער', 'תיאור']
    },
    'revenue_forecast': {
        'sheet_name': 'תחזית הכנסות',
        'table_headers': [
            'קומה', 'מס"ד', 'מספר עוקבי תכנית', 'כיוון', 'חדרים',
            'שטח פלוס במ"ר', 'מרפסת שמש', 'מרפסת גג/אחוד',
            'שווי למ"ר', 'שווי כללי', 'שווי כולל מע"מ'
        ]
    },
    'cost_forecast': {
        'sheet_name': 'תחזית עלויות',
        'table_headers': ['קטגוריה', 'כמות', 'יחידה', 'עלות ליחידה', 'סה"כ עלות']
    },
    'profitability': {
        'sheet_name': 'רווחיות',
        'table_headers': ['סיוג', 'אחוז', 'סכום (₪)']
    },
    'sensitivity_analysis': {
        'sheet_name': 'ניתוח רגישות',
        'table_headers': ['תרחיש', 'שינוי הכנסות', 'שינוי עלויות', 'שינוי רווח', 'אחוז רווח']
    },
    'break_even': {
        'sheet_name': 'נקודת איזון',
        'table_headers': ['פריט', 'ערך']
    },
    'land_value': {
        'sheet_name': 'ערך קרקע',
        'table_headers': ['תיאור', 'סכום למ"ר', 'סה"כ']
    },
    'index_values': {
        'sheet_name': 'מדדים',
        'fields': {
            'מדד תשומות בנייה': 'construction_cost_index',
            'מדד המחירים לצרכן': 'residential_index'
        }
    },
    'insurance': {
        'sheet_name': 'ביטוח',
        'table_headers': ['סוג ביטוח', 'כיסוי', 'עלות שנתית (₪)']
    },
    'cashflow': {
        'sheet_name': 'תזרים מזומנים',
        'table_headers': ['חודש', 'הכנסות (₪)', 'הוצאות (₪)', 'תזרים נקי (₪)', 'מצטבר (₪)']
    }
}

def create_excel_template():
    """Create Excel template with all sections"""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    field_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    field_font = Font(bold=True, size=11)
    
    for section_id, config in TEMPLATE_SECTIONS.items():
        ws = wb.create_sheet(title=config['sheet_name'])
        
        row = 1
        
        # Add fields section if exists
        if 'fields' in config:
            ws.cell(row=row, column=1, value="שדות:")
            ws.cell(row=row, column=1).font = Font(bold=True, size=14)
            row += 1
            
            for field_label in config['fields'].keys():
                ws.cell(row=row, column=1, value=field_label)
                ws.cell(row=row, column=1).fill = field_fill
                ws.cell(row=row, column=1).font = field_font
                ws.cell(row=row, column=2, value="")  # Empty cell for data
                row += 1
            
            row += 2  # Space before table
        
        # Add table headers if exists
        if 'table_headers' in config:
            ws.cell(row=row, column=1, value="טבלה:")
            ws.cell(row=row, column=1).font = Font(bold=True, size=14)
            row += 1
            
            # Create header row
            for col_idx, header in enumerate(config['table_headers'], start=1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                
                # Set column width
                ws.column_dimensions[get_column_letter(col_idx)].width = 20
            
            # Add 5 empty rows for data entry
            for _ in range(5):
                row += 1
        
        # Freeze first row
        ws.freeze_panes = 'A2'
    
    return wb

def generate_template_file(output_path):
    """Generate template file and save to path"""
    wb = create_excel_template()
    wb.save(output_path)
    return output_path
