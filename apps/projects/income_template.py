"""
Income Template Generator
Creates Excel template for apartment income data input
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json

def create_income_input_template(output_path):
    """Create Excel template for income data input"""
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create main sheet
    ws = wb.create_sheet(title="הכנסות - קלט נתונים")
    
    # Styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    section_font = Font(bold=True, size=14, color="2F5597")
    field_fill = PatternFill(start_color="E8F1FF", end_color="E8F1FF", fill_type="solid")
    
    row = 1
    
    # Constants Section
    ws.cell(row=row, column=1, value="קבועי חישוב")
    ws.cell(row=row, column=1).font = section_font
    row += 1
    
    # Constants fields
    constants_fields = [
        ("מחיר בסיס למ\"ר (₪)", "base_price_per_sqm", 32000),
        ("שיעור מע\"מ (%)", "vat_rate", 17),
        ("משקל מרפסת שמש (%)", "balcony_weight", 50),
        ("משקל מרפסת גג (%)", "roof_balcony_weight", 30),
        ("תוספת לקומה (%)", "floor_premium_per_floor", 1),
    ]
    
    for label, field_id, default_value in constants_fields:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=1).fill = field_fill
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=default_value)
        row += 1
    
    row += 1  # Space
    
    # Direction Premiums Section
    ws.cell(row=row, column=1, value="תוספות לכיוון (%)")
    ws.cell(row=row, column=1).font = section_font
    row += 1
    
    direction_premiums = [
        ("צפון", 0),
        ("דרום", 5),
        ("מזרח", 2),
        ("מערב", 2),
        ("צפון-מזרח", 3),
        ("צפון-מערב", 3),
        ("דרום-מזרח", 7),
        ("דרום-מערב", 7)
    ]
    
    for direction, premium in direction_premiums:
        ws.cell(row=row, column=1, value=direction)
        ws.cell(row=row, column=1).fill = field_fill
        ws.cell(row=row, column=2, value=premium)
        row += 1
    
    row += 2  # Space before apartments table
    
    # Apartments Table
    ws.cell(row=row, column=1, value="נתוני דירות")
    ws.cell(row=row, column=1).font = section_font
    row += 1
    
    # Table headers
    apartment_headers = [
        "בניין",
        "קומה", 
        "מס\"ד",
        "כיוון",
        "חדרים",
        "שטח דירה (מ\"ר)",
        "מרפסת שמש (מ\"ר)", 
        "מרפסת גג/חצר (מ\"ר)"
    ]
    
    for col, header in enumerate(apartment_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Sample data rows
    sample_apartments = [
        ["A", 0, 1, "צפון-מערב", "גן 4 חד'", 102.5, 14.0, 0],
        ["A", 1, 2, "דרום-מזרח", "5 חד'", 113.76, 14.05, 0],
        ["A", 1, 3, "צפון", "4 חד'", 95.0, 12.0, 0],
        ["A", 2, 4, "דרום", "3 חד'", 85.5, 10.0, 15.0],
        ["B", 0, 5, "מזרח", "גן 5 חד'", 120.0, 16.0, 25.0]
    ]
    
    for apt_data in sample_apartments:
        row += 1
        for col, value in enumerate(apt_data, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Add empty rows for more data
    for i in range(10):
        row += 1
        # Add empty row with just apartment number
        ws.cell(row=row, column=3, value=len(sample_apartments) + i + 1)
    
    # Set column widths
    column_widths = [8, 6, 6, 12, 12, 15, 18, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Add instructions sheet
    instructions_sheet = wb.create_sheet(title="הוראות שימוש")
    instructions = [
        "הוראות למילוי טופס חישוב הכנסות",
        "",
        "1. קבועי חישוב:",
        "   - מלא את המחיר הבסיסי למ\"ר",
        "   - הגדר את שיעור המע\"מ",
        "   - קבע משקלים למרפסות",
        "   - הגדר תוספת לקומה",
        "",
        "2. תוספות לכיוון:", 
        "   - ערוך את האחוזים לפי מדיניות התמחור",
        "",
        "3. נתוני דירות:",
        "   - מלא את כל הנתונים הנדרשים לכל דירה",
        "   - כיוון: בחר מהרשימה בסעיף 2",
        "   - שטחים: הכנס במ\"ר",
        "",
        "4. לאחר המילוי:",
        "   - העלה את הקובץ למערכת",
        "   - המערכת תחשב אוטומטית את כל המחירים",
        "   - תקבל דוח מפורט עם סיכומים"
    ]
    
    for i, instruction in enumerate(instructions, 1):
        cell = instructions_sheet.cell(row=i, column=1, value=instruction)
        if instruction and not instruction.startswith(" "):
            cell.font = Font(bold=True)
    
    wb.save(output_path)
    return output_path

def parse_income_template(file_path):
    """Parse filled income template and extract data"""
    from openpyxl import load_workbook
    
    wb = load_workbook(file_path, data_only=True)
    ws = wb["הכנסות - קלט נתונים"]
    
    # Parse constants
    constants = {
        'base_price_per_sqm': ws['B2'].value or 32000,
        'vat_rate': (ws['B3'].value or 17) / 100,
        'balcony_weight': (ws['B4'].value or 50) / 100, 
        'roof_balcony_weight': (ws['B5'].value or 30) / 100,
        'floor_premium_per_floor': (ws['B6'].value or 1) / 100
    }
    
    # Parse direction premiums
    direction_premiums = {}
    for row in range(8, 16):  # Direction premium rows
        direction = ws.cell(row=row, column=1).value
        premium = ws.cell(row=row, column=2).value
        if direction and premium is not None:
            direction_premiums[direction] = premium / 100
    
    constants['direction_premiums'] = direction_premiums
    
    # Parse apartments (starting from row with headers)
    apartments = []
    header_row = None
    
    # Find apartment table header
    for row in range(1, 50):
        if ws.cell(row=row, column=1).value == "בניין":
            header_row = row
            break
    
    if header_row:
        # Read apartment data
        for row in range(header_row + 1, ws.max_row + 1):
            building = ws.cell(row=row, column=1).value
            floor = ws.cell(row=row, column=2).value
            apt_num = ws.cell(row=row, column=3).value
            
            # Skip empty rows
            if not building and not apt_num:
                continue
                
            apartment = {
                'building': str(building) if building else 'A',
                'floor': int(floor) if floor is not None else 0,
                'apartment_num': int(apt_num) if apt_num else 0,
                'direction': ws.cell(row=row, column=4).value or 'צפון',
                'rooms': ws.cell(row=row, column=5).value or '3 חד\'',
                'apartment_area': float(ws.cell(row=row, column=6).value or 0),
                'sun_balcony': float(ws.cell(row=row, column=7).value or 0),
                'roof_balcony': float(ws.cell(row=row, column=8).value or 0)
            }
            
            # Only add if has minimum required data
            if apartment['apartment_area'] > 0:
                apartments.append(apartment)
    
    return {
        'constants': constants,
        'apartments': apartments
    }

if __name__ == "__main__":
    # Test template creation
    template_path = "/tmp/income_template.xlsx"
    create_income_input_template(template_path)
    print(f"✅ Template created: {template_path}")
    
    # Test parsing
    try:
        data = parse_income_template(template_path)
        print(f"✅ Parsed {len(data['apartments'])} apartments")
        print(f"✅ Base price: {data['constants']['base_price_per_sqm']:,} ₪")
    except Exception as e:
        print(f"❌ Parse error: {e}")
