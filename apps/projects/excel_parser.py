import pandas as pd
from openpyxl import load_workbook

def parse_uploaded_excel(file_path):
    """Parse uploaded Excel file and extract data from all sheets"""
    wb = load_workbook(file_path, data_only=True)
    parsed_data = {}
    
    # Sheet name to section ID mapping
    sheet_mapping = {
        'פרטי נכס': 'property_details',
        'תיאור פרויקט': 'project_description',
        'לוח זמנים': 'timeline',
        'לוח מכירות': 'sales_timeline',
        'תחזית הכנסות': 'revenue_forecast',
        'תחזית עלויות': 'cost_forecast',
        'רווחיות': 'profitability',
        'ניתוח רגישות': 'sensitivity_analysis',
        'נקודת איזון': 'break_even',
        'ערך קרקע': 'land_value',
        'מדדים': 'index_values',
        'ביטוח': 'insurance',
        'תזרים מזומנים': 'cashflow'
    }
    
    for sheet_name in wb.sheetnames:
        section_id = sheet_mapping.get(sheet_name)
        if not section_id:
            continue
            
        ws = wb[sheet_name]
        section_data = {}
        
        # Parse fields (key-value pairs)
        fields_started = False
        table_started = False
        table_data = []
        table_headers = []
        
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if not any(row):  # Skip empty rows
                continue
                
            # Check if this is "שדות:" marker
            if row[0] == "שדות:":
                fields_started = True
                table_started = False
                continue
            
            # Check if this is "טבלה:" marker
            if row[0] == "טבלה:":
                fields_started = False
                table_started = True
                continue
            
            # Parse fields section
            if fields_started and len(row) >= 2:
                field_name = row[0]
                field_value = row[1]
                if field_name and field_value:
                    # Convert Hebrew field name to English key
                    field_key = field_name.lower().replace(' ', '_')
                    section_data[field_key] = field_value
            
            # Parse table section
            if table_started:
                if not table_headers and row[0]:  # First row with data is headers
                    table_headers = [cell for cell in row if cell]
                    continue
                
                if table_headers and row[0]:  # Data row
                    row_dict = {}
                    for idx, header in enumerate(table_headers):
                        if idx < len(row) and row[idx]:
                            row_dict[header] = row[idx]
                    if row_dict:
                        table_data.append(row_dict)
        
        # Store parsed data
        if table_data:
            section_data['table_data'] = table_data
        
        parsed_data[section_id] = section_data
    
    return parsed_data

def validate_excel_structure(file_path):
    """Validate that uploaded Excel has correct structure"""
    wb = load_workbook(file_path, data_only=True)
    errors = []
    
    required_sheets = [
        'פרטי נכס', 'תיאור פרויקט', 'לוח זמנים', 'לוח מכירות',
        'תחזית הכנסות', 'תחזית עלויות', 'רווחיות', 'ניתוח רגישות',
        'נקודת איזון', 'ערך קרקע', 'מדדים', 'ביטוח', 'תזרים מזומנים'
    ]
    
    existing_sheets = wb.sheetnames
    
    for required in required_sheets:
        if required not in existing_sheets:
            errors.append(f"חסר גיליון: {required}")
    
    if errors:
        return False, errors
    
    return True, []
