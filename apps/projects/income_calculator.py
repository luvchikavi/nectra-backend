"""
Income Calculation Engine
Replicates the complex Excel logic for apartment revenue calculations
Based on analysis of income_master_file.xlsx
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import json

class IncomeCalculator:
    """
    Handles apartment revenue calculations with constants and business logic
    """
    
    def __init__(self, constants=None):
        """Initialize with project constants"""
        self.constants = constants or self.get_default_constants()
    
    def get_default_constants(self):
        """Default constants based on Excel analysis"""
        return {
            'base_price_per_sqm': 32000,     # C2 value from Excel
            'vat_rate': 0.17,                # C15 - 1 = 0.17 (17%)
            'balcony_weight': 0.5,           # C12 from Excel
            'roof_balcony_weight': 0.3,      # C13 from Excel
            'area_factor': 0.35,             # C14 from Excel
            'direction_premiums': {          # Based on VLOOKUP patterns
                'צפון': 0.0,
                'דרום': 0.05,
                'מזרח': 0.02,
                'מערב': 0.02,
                'צפון-מזרח': 0.03,
                'צפון-מערב': 0.03,
                'דרום-מזרח': 0.07,
                'דרום-מערב': 0.07
            },
            'floor_premium_per_floor': 0.01  # 1% per floor above ground
        }
    
    def calculate_apartment_revenue(self, apartments_data):
        """
        Calculate revenue for all apartments
        apartments_data: list of apartment dictionaries
        """
        results = []
        
        for apt in apartments_data:
            calculation = self.calculate_single_apartment(apt)
            results.append(calculation)
        
        return results
    
    def calculate_single_apartment(self, apartment):
        """Calculate revenue for single apartment - replicates Excel logic"""
        
        # Extract apartment data
        building = apartment.get('building', 'A')
        floor = int(apartment.get('floor', 0))
        apt_num = apartment.get('apartment_num', 0)
        direction = apartment.get('direction', 'צפון')
        rooms = apartment.get('rooms', '3 חד׳')
        apartment_area = float(apartment.get('apartment_area', 0))
        sun_balcony = float(apartment.get('sun_balcony', 0))
        roof_balcony = float(apartment.get('roof_balcony', 0))
        
        # Step 1: Calculate equivalent area (like Excel formula in column S)
        # Formula: =+L2+M2*$C$12 (apartment area + sun_balcony * balcony_weight)
        equivalent_area = apartment_area + (sun_balcony * self.constants['balcony_weight'])
        
        if roof_balcony > 0:
            # Add roof balcony with its weight factor
            equivalent_area += roof_balcony * self.constants['roof_balcony_weight']
        
        # Step 2: Direction premium (like Excel VLOOKUP in column T)
        direction_multiplier = 1.0 + self.constants['direction_premiums'].get(direction, 0.0)
        
        # Step 3: Floor premium (like Excel formula in column V)
        floor_multiplier = 1.0 + (max(0, floor - 1) * self.constants['floor_premium_per_floor'])
        
        # Step 4: Base price calculation (like Excel column P)
        base_total = equivalent_area * direction_multiplier * floor_multiplier
        
        # Step 5: Price per sqm (like Excel formula in column O)
        # Formula: =+ROUND(P2/S2,-1) - rounded to nearest 10
        price_per_sqm = round((base_total * self.constants['base_price_per_sqm'] / equivalent_area), -1)
        
        # Step 6: Total price without VAT (like Excel column Q)
        # Formula: =(P2/$C$15) 
        total_without_vat = round(equivalent_area * price_per_sqm)
        
        # Step 7: Total price with VAT (like Excel column P)
        total_with_vat = round(total_without_vat * (1 + self.constants['vat_rate']))
        
        return {
            'building': building,
            'floor': floor,
            'apartment_num': apt_num,
            'direction': direction,
            'rooms': rooms,
            'apartment_area': apartment_area,
            'sun_balcony': sun_balcony,
            'roof_balcony': roof_balcony,
            'equivalent_area': round(equivalent_area, 2),
            'direction_multiplier': direction_multiplier,
            'floor_multiplier': floor_multiplier,
            'price_per_sqm': int(price_per_sqm),
            'total_without_vat': int(total_without_vat),
            'total_with_vat': int(total_with_vat),
            'total_area_for_calculation': round(equivalent_area, 2)
        }
    
    def generate_summary(self, calculated_apartments):
        """Generate project summary statistics"""
        if not calculated_apartments:
            return {}
        
        total_units = len(calculated_apartments)
        total_area = sum(apt['apartment_area'] for apt in calculated_apartments)
        total_equivalent_area = sum(apt['equivalent_area'] for apt in calculated_apartments)
        total_revenue_no_vat = sum(apt['total_without_vat'] for apt in calculated_apartments)
        total_revenue_with_vat = sum(apt['total_with_vat'] for apt in calculated_apartments)
        
        avg_price_per_sqm = total_revenue_no_vat / total_equivalent_area if total_equivalent_area > 0 else 0
        avg_apartment_price = total_revenue_no_vat / total_units if total_units > 0 else 0
        
        return {
            'total_units': total_units,
            'total_apartment_area': round(total_area, 2),
            'total_equivalent_area': round(total_equivalent_area, 2),
            'total_revenue_without_vat': int(total_revenue_no_vat),
            'total_revenue_with_vat': int(total_revenue_with_vat),
            'average_price_per_sqm': round(avg_price_per_sqm, 0),
            'average_apartment_price': round(avg_apartment_price, 0),
            'total_vat_amount': int(total_revenue_with_vat - total_revenue_no_vat)
        }

def create_income_excel_export(apartments_data, constants, output_path):
    """Create professional Excel export with calculations"""
    
    calculator = IncomeCalculator(constants)
    calculated_apartments = calculator.calculate_apartment_revenue(apartments_data)
    summary = calculator.generate_summary(calculated_apartments)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "חישוב הכנסות"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    number_format = '#,##0'
    currency_format = '#,##0 ₪'
    
    # Constants section
    ws['A1'] = "קבועי חישוב"
    ws['A1'].font = Font(bold=True, size=14)
    
    constants_data = [
        ['מחיר בסיס למ"ר', constants['base_price_per_sqm'], '₪'],
        ['שיעור מע"מ', f"{constants['vat_rate']*100:.0f}%", ''],
        ['משקל מרפסת שמש', f"{constants['balcony_weight']*100:.0f}%", ''],
        ['תוספת לקומה', f"{constants['floor_premium_per_floor']*100:.0f}%", 'לקומה']
    ]
    
    for i, (label, value, unit) in enumerate(constants_data, 2):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'C{i}'] = unit
    
    # Main table
    start_row = 8
    headers = [
        'בניין', 'קומה', 'מס"ד', 'כיוון', 'חדרים',
        'שטח דירה', 'מרפסת שמש', 'מרפסת גג',
        'שטח מקביל', 'מחיר למ"ר', 'סה"כ ללא מע"מ', 'סה"כ כולל מע"מ'
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
    
    # Write apartment data
    for row_idx, apt in enumerate(calculated_apartments, start_row + 1):
        row_data = [
            apt['building'],
            apt['floor'],
            apt['apartment_num'],
            apt['direction'],
            apt['rooms'],
            apt['apartment_area'],
            apt['sun_balcony'],
            apt['roof_balcony'],
            apt['equivalent_area'],
            apt['price_per_sqm'],
            apt['total_without_vat'],
            apt['total_with_vat']
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.value = value
            
            # Format numbers
            if col >= 6:  # Numeric columns
                cell.number_format = currency_format if col >= 10 else number_format
    
    # Summary section
    summary_start = start_row + len(calculated_apartments) + 3
    ws[f'A{summary_start}'] = "סיכום פרויקט"
    ws[f'A{summary_start}'].font = Font(bold=True, size=14)
    
    summary_data = [
        ['סה"כ יחידות', summary['total_units']],
        ['סה"כ שטח דירות', f"{summary['total_apartment_area']:.0f} מ\"ר"],
        ['סה"כ הכנסות ללא מע"מ', f"{summary['total_revenue_without_vat']:,} ₪"],
        ['סה"כ הכנסות כולל מע"מ', f"{summary['total_revenue_with_vat']:,} ₪"],
        ['מחיר ממוצע למ"ר', f"{summary['average_price_per_sqm']:,.0f} ₪"],
        ['מחיר ממוצע לדירה', f"{summary['average_apartment_price']:,} ₪"]
    ]
    
    for i, (label, value) in enumerate(summary_data, summary_start + 1):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'A{i}'].font = Font(bold=True)
    
    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_path)
    return output_path

# Test function
def test_income_calculator():
    """Test the calculator with sample data"""
    
    sample_apartments = [
        {
            'building': 'A',
            'floor': 0,
            'apartment_num': 1,
            'direction': 'צפון-מערב',
            'rooms': 'גן 4 חד׳',
            'apartment_area': 102.5,
            'sun_balcony': 14.0,
            'roof_balcony': 0
        },
        {
            'building': 'A', 
            'floor': 1,
            'apartment_num': 2,
            'direction': 'דרום-מזרח',
            'rooms': '5 חד׳',
            'apartment_area': 113.76,
            'sun_balcony': 14.05,
            'roof_balcony': 0
        }
    ]
    
    calculator = IncomeCalculator()
    results = calculator.calculate_apartment_revenue(sample_apartments)
    summary = calculator.generate_summary(results)
    
    print("✅ TEST RESULTS:")
    for i, result in enumerate(results):
        print(f"Apartment {i+1}: {result['total_with_vat']:,} ₪")
    
    print(f"Total Revenue: {summary['total_revenue_with_vat']:,} ₪")
    return True

if __name__ == "__main__":
    test_income_calculator()
